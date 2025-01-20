import openpyxl

wb = openpyxl.load_workbook("paquetes-populares/marcas_modelos.xlsx")

print(wb.sheetnames)

marcas = wb["marcas"]
# print(marcas.max_row, marcas.max_column)
# print(marcas["A1"].value)
# print(marcas.cell(row=1, column=2).value)

# agregar hoja
wb.create_sheet("Hoja 3")


# for fila in range(1, marcas.max_row + 1):
#     for columna in range(1, marcas.max_column + 1):
#         celda = marcas.cell(row=fila, column=columna).value
#         print(celda)


marcas.append([80, "new marca"])
marcas.delete_rows(4, 1)

wb.save("paquetes-populares/excel_modificado.xlsx")
